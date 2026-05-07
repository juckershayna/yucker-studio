from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

wb = Workbook()

# ── Farben ──────────────────────────────────────────────────────────────────
BLACK      = "FF1A1A1A"
DARK_GREY  = "FF2D2D2D"
MID_GREY   = "FF3D3D3D"
LIGHT_GREY = "FFF2F2F2"
WHITE      = "FFFFFFFF"
ACCENT     = "FFCFFF4F"   # Yucker lime-green
ACCENT2    = "FF8FBF00"
ROW_ALT    = "FFF7F7F7"
BORDER_COL = "FFD0D0D0"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=10, color=BLACK, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def border_thin():
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    s = Side(style="thin", color=BORDER_COL)
    return Border(bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right():
    return Alignment(horizontal="right", vertical="center")

CHF = '_-"CHF"\ * #\'##0_-;-"CHF"\ * #\'##0_-;_-"CHF"\ * "-"_-;_-@_-'
CHF_DEC = '_-"CHF"\ * #\'##0.00_-;-"CHF"\ * #\'##0.00_-;_-"CHF"\ * "-"??_-;_-@_-'
PCT = '0.0"%"'
NUM = '#\'##0'

def header_row(ws, row, cols, texts, bg=DARK_GREY, fg=WHITE, size=10):
    for col, text in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=text)
        c.font = font(bold=True, size=size, color=fg)
        c.fill = fill(bg)
        c.alignment = center()
        c.border = border_thin()

def data_row(ws, row, data, formats=None, bgs=None, bold_cols=None):
    for i, (col, val) in enumerate(data):
        c = ws.cell(row=row, column=col, value=val)
        bg = bgs[i] if bgs else WHITE
        c.fill = fill(bg)
        c.border = border_thin()
        if formats and formats[i]:
            c.number_format = formats[i]
        if bold_cols and col in bold_cols:
            c.font = font(bold=True)
        else:
            c.font = font()
        if isinstance(val, (int, float)):
            c.alignment = right()
        else:
            c.alignment = left()

def title_block(ws, row, col, title, subtitle=None):
    c = ws.cell(row=row, column=col, value=title)
    c.font = font(bold=True, size=14, color=BLACK)
    c.alignment = left()
    if subtitle:
        c2 = ws.cell(row=row+1, column=col, value=subtitle)
        c2.font = font(size=9, color="FF666666", italic=True)
        c2.alignment = left()

def section_header(ws, row, col_start, col_end, text):
    ws.cell(row=row, column=col_start, value=text)
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start)
    c.font = font(bold=True, size=10, color=WHITE)
    c.fill = fill(MID_GREY)
    c.alignment = left()
    c.border = border_thin()

def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(size=8, color="FF888888", italic=True)
    c.alignment = left()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – ÜBERSICHT / DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Übersicht"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 2
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["C"].width = 18
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 18
ws1.column_dimensions["F"].width = 18
ws1.column_dimensions["G"].width = 2

# Title
ws1.row_dimensions[1].height = 10
ws1.row_dimensions[2].height = 36
ws1.row_dimensions[3].height = 18
for col in range(1, 8):
    ws1.cell(row=2, column=col).fill = fill(BLACK)
    ws1.cell(row=3, column=col).fill = fill(BLACK)

c = ws1.cell(row=2, column=2, value="YUCKER STUDIO — Finanzübersicht 2026")
c.font = Font(name="Calibri", bold=True, size=16, color=ACCENT[2:])
c.alignment = Alignment(horizontal="left", vertical="center")
c = ws1.cell(row=3, column=2, value="Kanton Zürich · 3 Personen · Zielstundensatz CHF 140/h")
c.font = Font(name="Calibri", size=9, color="FFAAAAAA")
c.alignment = left()

# ── KOSTEN ──────────────────────────────────────────────────────────────────
r = 5
section_header(ws1, r, 2, 6, "  MONATLICHE BETRIEBSKOSTEN")
r += 1
header_row(ws1, r, [2,3,4,5,6],
           ["Position","CHF / Monat","CHF / Jahr","Anteil","Notiz"],
           bg=DARK_GREY)
r += 1

costs = [
    ("Büro / Coworking Zürich",          2500,  "Schätzung Zürich-Stadt"),
    ("Lohn Inhaber 1 (brutto)",           8000,  "Angemessen für GmbH/AG Inhaber"),
    ("Lohn Inhaber 2 (brutto)",           8000,  "Angemessen für GmbH/AG Inhaber"),
    ("Lohn Mitarbeiterin (brutto)",        6000,  "Kreativbranche Zürich ~CHF 72k/J"),
    ("AHV / UV / KTG / BVG (~27%)",       5940,  "Auf CHF 22'000 Lohnsumme"),
    ("Software & Lizenzen",                600,  "Adobe CC, Figma, etc."),
    ("Ausrüstung (Amortisation)",          500,  "Kamera, Computer, etc."),
    ("Buchhaltung & Steuerberatung",       400,  ""),
    ("Versicherungen",                     300,  ""),
    ("Marketing & Diverses",               400,  "Website, Visitenkarten, etc."),
]
total_monthly = sum(x[1] for x in costs)
total_yearly  = total_monthly * 12

for i, (pos, mo, notiz) in enumerate(costs):
    bg = ROW_ALT if i % 2 == 0 else WHITE
    pct = mo / total_monthly * 100
    ws1.cell(row=r, column=2, value=pos).fill = fill(bg)
    ws1.cell(row=r, column=2).font = font()
    ws1.cell(row=r, column=2).alignment = left()
    ws1.cell(row=r, column=2).border = border_thin()

    ws1.cell(row=r, column=3, value=mo).fill = fill(bg)
    ws1.cell(row=r, column=3).number_format = CHF
    ws1.cell(row=r, column=3).alignment = right()
    ws1.cell(row=r, column=3).border = border_thin()
    ws1.cell(row=r, column=3).font = font()

    ws1.cell(row=r, column=4, value=mo*12).fill = fill(bg)
    ws1.cell(row=r, column=4).number_format = CHF
    ws1.cell(row=r, column=4).alignment = right()
    ws1.cell(row=r, column=4).border = border_thin()
    ws1.cell(row=r, column=4).font = font()

    ws1.cell(row=r, column=5, value=pct/100).fill = fill(bg)
    ws1.cell(row=r, column=5).number_format = '0.0%'
    ws1.cell(row=r, column=5).alignment = right()
    ws1.cell(row=r, column=5).border = border_thin()
    ws1.cell(row=r, column=5).font = font()

    ws1.cell(row=r, column=6, value=notiz).fill = fill(bg)
    ws1.cell(row=r, column=6).font = font(size=8, italic=True, color="FF777777")
    ws1.cell(row=r, column=6).alignment = left()
    ws1.cell(row=r, column=6).border = border_thin()
    r += 1

# Total row
for col in [2,3,4,5,6]:
    ws1.cell(row=r, column=col).fill = fill(ACCENT)
    ws1.cell(row=r, column=col).border = border_thin()
ws1.cell(row=r, column=2, value="TOTAL MONATLICHE KOSTEN").font = font(bold=True)
ws1.cell(row=r, column=2).alignment = left()
ws1.cell(row=r, column=3, value=total_monthly).number_format = CHF
ws1.cell(row=r, column=3).font = font(bold=True)
ws1.cell(row=r, column=3).alignment = right()
ws1.cell(row=r, column=4, value=total_yearly).number_format = CHF
ws1.cell(row=r, column=4).font = font(bold=True)
ws1.cell(row=r, column=4).alignment = right()
ws1.cell(row=r, column=5, value=1.0).number_format = '0.0%'
ws1.cell(row=r, column=5).font = font(bold=True)
ws1.cell(row=r, column=5).alignment = right()
r += 2

# ── STUNDENSATZ ─────────────────────────────────────────────────────────────
section_header(ws1, r, 2, 6, "  STUNDENSATZ-KALKULATION")
r += 1
header_row(ws1, r, [2,3,4,5,6],
           ["Kennzahl","Wert","Einheit","","Notiz"],
           bg=DARK_GREY)
r += 1

kpi_rows = [
    ("Anzahl Mitarbeitende", 3, "Personen", "", "2 Inhaber + 1 Mitarbeiterin"),
    ("Arbeitsstunden / Person / Monat", 160, "h", "", "Vollzeit ~40h/Woche"),
    ("Theoretische Stunden total / Monat", 480, "h", "", "3 × 160h"),
    ("Fakturierbare Stunden (55%)", 264, "h", "", "Admin, Akquise, unbezahlte Zeit abgezogen"),
    ("Break-Even-Stundensatz", round(total_monthly/264), "CHF/h", "", f"CHF {total_monthly:,} / 264h"),
    ("Zielstundensatz (mit 20% Marge)", round(total_monthly/264*1.20), "CHF/h", "", "Empfohlener Preis"),
    ("Angestrebter Stundensatz", 140, "CHF/h", "", "Gewählt: CHF 140/h (Marktniveau Zürich)"),
    ("Monatlicher Umsatz (Ziel)", 140*264, "CHF/Monat", "", "264h × CHF 140/h"),
    ("Jahresumsatz (Ziel)", 140*264*12, "CHF/Jahr", "", ""),
    ("Gewinn vor Steuern (Ziel)", 140*264*12 - total_yearly, "CHF/Jahr", "", ""),
]
for i, (k, v, u, e, n) in enumerate(kpi_rows):
    bg = ROW_ALT if i % 2 == 0 else WHITE
    is_highlight = k in ("Angestrebter Stundensatz", "Monatlicher Umsatz (Ziel)")
    bg_use = ACCENT if is_highlight else bg
    bold = is_highlight

    ws1.cell(row=r, column=2, value=k).fill = fill(bg_use)
    ws1.cell(row=r, column=2).font = font(bold=bold)
    ws1.cell(row=r, column=2).alignment = left()
    ws1.cell(row=r, column=2).border = border_thin()

    c3 = ws1.cell(row=r, column=3, value=v)
    c3.fill = fill(bg_use)
    c3.font = font(bold=bold)
    c3.alignment = right()
    c3.border = border_thin()
    if "CHF" in u:
        c3.number_format = CHF
    else:
        c3.number_format = NUM

    ws1.cell(row=r, column=4, value=u).fill = fill(bg_use)
    ws1.cell(row=r, column=4).font = font(bold=bold)
    ws1.cell(row=r, column=4).alignment = left()
    ws1.cell(row=r, column=4).border = border_thin()

    ws1.cell(row=r, column=5, value="").fill = fill(bg_use)
    ws1.cell(row=r, column=5).border = border_thin()

    ws1.cell(row=r, column=6, value=n).fill = fill(bg_use)
    ws1.cell(row=r, column=6).font = font(size=8, italic=True, color="FF777777")
    ws1.cell(row=r, column=6).alignment = left()
    ws1.cell(row=r, column=6).border = border_thin()
    r += 1

r += 1
note(ws1, r, 2, "Quelle: Eigene Kalkulation basierend auf Zürcher Marktdaten 2026. Lohnstrukturerhebung BFS, Branchenvergleiche Kreativwirtschaft Kanton Zürich.")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – LÖHNE
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Löhne & Personal")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 26
ws2.column_dimensions["C"].width = 16
ws2.column_dimensions["D"].width = 16
ws2.column_dimensions["E"].width = 16
ws2.column_dimensions["F"].width = 16
ws2.column_dimensions["G"].width = 16
ws2.column_dimensions["H"].width = 16
ws2.column_dimensions["I"].width = 2

for col in range(1, 10):
    ws2.cell(row=2, column=col).fill = fill(BLACK)
    ws2.cell(row=3, column=col).fill = fill(BLACK)
ws2.row_dimensions[2].height = 36
ws2.row_dimensions[3].height = 18
c = ws2.cell(row=2, column=2, value="LÖHNE & PERSONALKOSTEN")
c.font = Font(name="Calibri", bold=True, size=16, color=ACCENT[2:])
c.alignment = Alignment(horizontal="left", vertical="center")
c = ws2.cell(row=3, column=2, value="Yucker Studio · Kanton Zürich 2026")
c.font = Font(name="Calibri", size=9, color="FFAAAAAA")
c.alignment = left()

r = 5
section_header(ws2, r, 2, 8, "  LOHNÜBERSICHT PRO PERSON")
r += 1
header_row(ws2, r, [2,3,4,5,6,7,8],
           ["Person / Rolle","Bruttolohn/Mo","Bruttolohn/J","Stundenlohn*",
            "Sozialabgaben/J\n(~27%)","Gesamtkosten/J","Gesamtkosten/Mo"],
           bg=DARK_GREY)
r += 1

persons = [
    ("Inhaber 1",       8000,  "2 Inhaber, Eigenverantwortung, Kundenkontakt"),
    ("Inhaber 2",       8000,  ""),
    ("Mitarbeiterin",   6000,  "Kreativbranche Zürich. Median ~CHF 70k/J"),
]
total_lohn_mo = sum(p[1] for p in persons)
total_lohn_j  = total_lohn_mo * 12
total_sv_j    = round(total_lohn_j * 0.27)
total_pers_j  = total_lohn_j + total_sv_j

for i, (name, mo, notiz) in enumerate(persons):
    bg = ROW_ALT if i % 2 == 0 else WHITE
    j = mo * 12
    sv = round(j * 0.27)
    gesamt_j = j + sv
    gesamt_mo = round(gesamt_j / 12)
    stundenl = round(j / 1920, 2)   # 1920h = 48 Wochen × 40h

    for col in [2,3,4,5,6,7,8]:
        ws2.cell(row=r, column=col).fill = fill(bg)
        ws2.cell(row=r, column=col).border = border_thin()
        ws2.cell(row=r, column=col).font = font()

    ws2.cell(row=r, column=2, value=f"{name}").alignment = left()
    ws2.cell(row=r, column=2).font = font(bold=True)
    ws2.cell(row=r, column=3, value=mo).number_format = CHF
    ws2.cell(row=r, column=3).alignment = right()
    ws2.cell(row=r, column=4, value=j).number_format = CHF
    ws2.cell(row=r, column=4).alignment = right()
    ws2.cell(row=r, column=5, value=stundenl).number_format = CHF_DEC
    ws2.cell(row=r, column=5).alignment = right()
    ws2.cell(row=r, column=6, value=sv).number_format = CHF
    ws2.cell(row=r, column=6).alignment = right()
    ws2.cell(row=r, column=7, value=gesamt_j).number_format = CHF
    ws2.cell(row=r, column=7).alignment = right()
    ws2.cell(row=r, column=8, value=gesamt_mo).number_format = CHF
    ws2.cell(row=r, column=8).alignment = right()
    r += 1

# Total
for col in [2,3,4,5,6,7,8]:
    ws2.cell(row=r, column=col).fill = fill(ACCENT)
    ws2.cell(row=r, column=col).border = border_thin()
    ws2.cell(row=r, column=col).font = font(bold=True)

avg_stl = round((total_lohn_j / 3) / 1920, 2)
ws2.cell(row=r, column=2, value="TOTAL / DURCHSCHNITT").alignment = left()
ws2.cell(row=r, column=3, value=total_lohn_mo).number_format = CHF; ws2.cell(row=r, column=3).alignment = right()
ws2.cell(row=r, column=4, value=total_lohn_j).number_format = CHF; ws2.cell(row=r, column=4).alignment = right()
ws2.cell(row=r, column=5, value=avg_stl).number_format = CHF_DEC; ws2.cell(row=r, column=5).alignment = right()
ws2.cell(row=r, column=6, value=total_sv_j).number_format = CHF; ws2.cell(row=r, column=6).alignment = right()
ws2.cell(row=r, column=7, value=total_pers_j).number_format = CHF; ws2.cell(row=r, column=7).alignment = right()
ws2.cell(row=r, column=8, value=round(total_pers_j/12)).number_format = CHF; ws2.cell(row=r, column=8).alignment = right()
r += 2

note(ws2, r, 2, "* Bruttostundenlohn berechnet auf 1'920 Stunden/Jahr (48 Wochen × 40h). Sozialabgaben inkl. AHV/IV/EO, ALV, UVG, KTG, BVG (Arbeitgeberanteil).")
r += 2

# Vergleich Markt
section_header(ws2, r, 2, 8, "  MARKTVERGLEICH KREATIVBRANCHE ZÜRICH 2026")
r += 1
header_row(ws2, r, [2,3,4,5,6,7,8],
           ["Funktion","Min. Lohn/J","Median Lohn/J","Max. Lohn/J",
            "Min. CHF/h","Median CHF/h","Max. CHF/h"],
           bg=DARK_GREY)
r += 1
markt = [
    ("Grafiker/in (Junior)", 58000, 68000, 78000),
    ("Grafiker/in (Senior)", 72000, 84000, 98000),
    ("Web Designer/in",      72000, 88000, 105000),
    ("Fotograf/in",          60000, 72000, 90000),
    ("Marketing Manager/in", 70000, 85000, 105000),
    ("Interior Designer/in", 65000, 78000, 95000),
    ("Videograf/in",         62000, 75000, 92000),
    ("Studio-Inhaber/in",    85000, 100000, 130000),
]
for i, (fn, lo, med, hi) in enumerate(markt):
    bg = ROW_ALT if i % 2 == 0 else WHITE
    for col in [2,3,4,5,6,7,8]:
        ws2.cell(row=r, column=col).fill = fill(bg)
        ws2.cell(row=r, column=col).border = border_thin()
        ws2.cell(row=r, column=col).font = font()
    ws2.cell(row=r, column=2, value=fn).alignment = left()
    ws2.cell(row=r, column=3, value=lo).number_format = CHF; ws2.cell(row=r, column=3).alignment = right()
    ws2.cell(row=r, column=4, value=med).number_format = CHF; ws2.cell(row=r, column=4).alignment = right()
    ws2.cell(row=r, column=5, value=hi).number_format = CHF; ws2.cell(row=r, column=5).alignment = right()
    ws2.cell(row=r, column=6, value=round(lo/1920,0)).number_format = CHF; ws2.cell(row=r, column=6).alignment = right()
    ws2.cell(row=r, column=7, value=round(med/1920,0)).number_format = CHF; ws2.cell(row=r, column=7).alignment = right()
    ws2.cell(row=r, column=8, value=round(hi/1920,0)).number_format = CHF; ws2.cell(row=r, column=8).alignment = right()
    r += 1
r += 1
note(ws2, r, 2, "Quelle: Lohnstrukturerhebung BFS 2024, Jobcloud Salärindex 2025, eigene Recherche. Werte gerundet.")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – PREISE & STUNDENSÄTZE
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Preise & Stundensätze")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 2
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 16
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 14
ws3.column_dimensions["G"].width = 14
ws3.column_dimensions["H"].width = 22
ws3.column_dimensions["I"].width = 2

for col in range(1, 10):
    ws3.cell(row=2, column=col).fill = fill(BLACK)
    ws3.cell(row=3, column=col).fill = fill(BLACK)
ws3.row_dimensions[2].height = 36
ws3.row_dimensions[3].height = 18
c = ws3.cell(row=2, column=2, value="PREISLISTE & IMPLIZITER STUNDENSATZ")
c.font = Font(name="Calibri", bold=True, size=16, color=ACCENT[2:])
c.alignment = Alignment(horizontal="left", vertical="center")
c = ws3.cell(row=3, column=2, value="Alle Preise exkl. MwSt. · Zielstundensatz CHF 140/h · Stand 2026")
c.font = Font(name="Calibri", size=9, color="FFAAAAAA")
c.alignment = left()

r = 5

services = [
    ("FOTOGRAFIE", [
        ("1 Stunde",                   390,   1,    1,   "1h Shooting + Grundbearbeitung"),
        ("2–3 Stunden",                750,   2.5,  5,   "Avg 2.5h + 2.5h Nachbearbeitung"),
        ("Halbtag (4h)",              1100,   4,    7,   "4h + 3h Nachbearbeitung"),
        ("Ganztag (8h)",              1900,   8,   13,   "8h + 5h Nachbearbeitung"),
        ("  + Erweitert (60 Fotos)",   190,   None, 1.5, "Add-on"),
        ("  + Vollgalerie (100+ Fotos)",450,  None, 3,   "Add-on"),
    ]),
    ("FILM / VIDEO", [
        ("Social Reel 30s",           1100,   None, 7,   "Prep + Dreh + Schnitt"),
        ("Kurzfilm 60s",              1600,   None,11,   ""),
        ("Markenfilm 2–3 min",        2800,   None,19,   ""),
        ("Feature 3–5 min",           3900,   None,27,   ""),
        ("  + Drohnenaufnahmen",        650,  None, 4,   "Add-on"),
        ("  + Voice-Over",              380,  None, 2.5, "Add-on"),
        ("  + Soundtrack-Lizenz",       260,  None, 1.5, "Add-on"),
    ]),
    ("BRANDING", [
        ("Logo only",                 1600,   None,11,   "Research, Konzepte, Finalisierung"),
        ("Logo + Richtlinien",        2800,   None,20,   ""),
        ("Full Brand Identity",       4800,   None,34,   "Logo, Guidelines, Collateral"),
        ("Rebranding",                3900,   None,28,   ""),
        ("  + Visitenkarten",           320,  None, 2.5, "Add-on"),
        ("  + Briefpapier-Set",         550,  None, 4,   "Add-on"),
        ("  + Social-Media-Kit",        480,  None, 3.5, "Add-on"),
        ("  + Markenbuch",              850,  None, 6,   "Add-on"),
    ]),
    ("WEBDESIGN", [
        ("1–3 Seiten",                2300,   None,16,   "Design + Entwicklung + Deploy"),
        ("4–7 Seiten",                3900,   None,28,   ""),
        ("8–15 Seiten",               5900,   None,42,   ""),
        ("  + E-Commerce Shop",       1900,   None,14,   "Add-on"),
        ("  + Blog",                    550,  None, 4,   "Add-on"),
        ("  + Buchungssystem",          800,  None, 6,   "Add-on"),
        ("  + Mehrsprachig",            700,  None, 5,   "Add-on"),
    ]),
    ("SOCIAL MEDIA MARKETING", [
        ("Strategie & Konzept",        900,   None, 6,   "Einmalig"),
        ("Content-Erstellung",        1500,   None,11,   "Pro Monat / 1 Kanal"),
        ("Community-Management",      1100,   None, 8,   "Pro Monat / 1 Kanal"),
        ("Vollservice",               2500,   None,18,   "Pro Monat / 1 Kanal"),
    ]),
    ("INTERIOR DESIGN", [
        ("Beratung (2h)",              380,   2,    2.5, "CHF 190/h"),
        ("Konzept & Moodboard",       1100,   None, 8,   ""),
        ("Raumplanung",               1900,   None,14,   "Basis bis 30m²"),
        ("Vollprojekt",               3800,   None,27,   "Basis bis 30m²"),
        ("  + 3D-Visualisierung",       850,  None, 6,   "Add-on"),
        ("  + Einkaufsliste",           280,  None, 2,   "Add-on"),
        ("  + Handwerkerkoordination",  700,  None, 5,   "Add-on"),
    ]),
]

header_row(ws3, r, [2,3,4,5,6,7,8],
           ["Leistung","Preis (CHF)","Preis inkl.\nRush +40%",
            "Typ. Stunden\n(geschätzt)","Impl. Stundensatz",
            "Markt Zürich\nStundensatz","Bemerkung"],
           bg=DARK_GREY)
r += 1

target_h = 140
for svc_name, items in services:
    # Section bar
    section_header(ws3, r, 2, 8, f"  {svc_name}")
    r += 1
    for i, row_data in enumerate(items):
        name, price, _, hours, remark = row_data
        is_addon = name.strip().startswith("+")
        bg = "FFF5FFF0" if not is_addon else WHITE

        impl_h = round(price / hours) if hours else None
        rush = round(price * 1.4) if not is_addon else None
        markt_low, markt_high = 130, 160

        for col in [2,3,4,5,6,7,8]:
            ws3.cell(row=r, column=col).fill = fill(bg)
            ws3.cell(row=r, column=col).border = border_thin()

        ws3.cell(row=r, column=2, value=name).font = font(bold=not is_addon, size=9 if is_addon else 10)
        ws3.cell(row=r, column=2).alignment = left()

        ws3.cell(row=r, column=3, value=price).number_format = CHF
        ws3.cell(row=r, column=3).alignment = right()
        ws3.cell(row=r, column=3).font = font(bold=not is_addon)

        ws3.cell(row=r, column=4, value=rush).number_format = CHF if rush else "@"
        ws3.cell(row=r, column=4).alignment = right()
        ws3.cell(row=r, column=4).font = font(size=9, color="FF555555")

        ws3.cell(row=r, column=5, value=f"~{hours}h" if hours else "–").alignment = right()
        ws3.cell(row=r, column=5).font = font(size=9, color="FF555555")

        if impl_h:
            color = "FF008000" if impl_h >= 130 else ("FFCC6600" if impl_h >= 100 else "FFCC0000")
            ws3.cell(row=r, column=6, value=impl_h).number_format = CHF
            ws3.cell(row=r, column=6).alignment = right()
            ws3.cell(row=r, column=6).font = Font(name="Calibri", bold=True, size=10, color=color)
        else:
            ws3.cell(row=r, column=6, value="Add-on").alignment = right()
            ws3.cell(row=r, column=6).font = font(size=9, color="FFAAAAAA")

        ws3.cell(row=r, column=7, value="CHF 130–160/h").alignment = right()
        ws3.cell(row=r, column=7).font = font(size=9, color="FF555555")

        ws3.cell(row=r, column=8, value=remark).alignment = left()
        ws3.cell(row=r, column=8).font = font(size=8, italic=True, color="FF777777")

        r += 1

r += 1
note(ws3, r, 2, "Grün = Stundensatz ≥ CHF 130/h ✓  |  Orange = CHF 100–129/h (akzeptabel)  |  Rot = unter CHF 100/h")
r += 1
note(ws3, r, 2, "Alle Preise exkl. 8.1% MWST. Endangebot nach Erstgespräch. Quelle: Eigene Kalkulation auf Basis Zürcher Kreativmarkt 2026.")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 – JAHRESPLANUNG
# ════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Jahresplanung")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 2
ws4.column_dimensions["B"].width = 18
for col in range(3, 16):
    ws4.column_dimensions[get_column_letter(col)].width = 10
ws4.column_dimensions[get_column_letter(16)].width = 12
ws4.column_dimensions[get_column_letter(17)].width = 12

for col in range(1, 18):
    ws4.cell(row=2, column=col).fill = fill(BLACK)
    ws4.cell(row=3, column=col).fill = fill(BLACK)
ws4.row_dimensions[2].height = 36
ws4.row_dimensions[3].height = 18
c = ws4.cell(row=2, column=2, value="JAHRESPLANUNG & BUDGET 2026")
c.font = Font(name="Calibri", bold=True, size=16, color=ACCENT[2:])
c.alignment = Alignment(horizontal="left", vertical="center")
c = ws4.cell(row=3, column=2, value="Monatliche Übersicht · Zielstundensatz CHF 140/h · 264 Fakt.-Stunden/Monat")
c.font = Font(name="Calibri", size=9, color="FFAAAAAA")
c.alignment = left()

r = 5
months = ["Jan","Feb","Mär","Apr","Mai","Jun","Jul","Aug","Sep","Okt","Nov","Dez","TOTAL","Ø/Monat"]
header_row(ws4, r, range(2, 18), [""] + months, bg=DARK_GREY, size=9)
r += 1

plan_rows = [
    ("Ziel-Umsatz",        36960, True),
    ("Kosten (fix)",       total_monthly, False),
    ("Betriebsergebnis",   36960 - total_monthly, True),
    ("Fakturstunden",      264, False),
    ("Ø Stundensatz",      140, False),
]

# Simple: all months equal for planning purposes
for i, (label, mo_val, is_bold) in enumerate(plan_rows):
    bg = ROW_ALT if i % 2 == 0 else WHITE
    is_highlight = label == "Betriebsergebnis"
    bg_use = ACCENT if is_highlight else bg

    ws4.cell(row=r, column=2, value=label).fill = fill(bg_use)
    ws4.cell(row=r, column=2).font = font(bold=is_bold)
    ws4.cell(row=r, column=2).alignment = left()
    ws4.cell(row=r, column=2).border = border_thin()

    yearly = mo_val * 12
    for m_idx, col in enumerate(range(3, 15)):  # Jan–Dez
        ws4.cell(row=r, column=col, value=mo_val).fill = fill(bg_use)
        ws4.cell(row=r, column=col).font = font(bold=is_highlight)
        ws4.cell(row=r, column=col).alignment = right()
        ws4.cell(row=r, column=col).border = border_thin()
        if label in ("Ziel-Umsatz","Kosten (fix)","Betriebsergebnis"):
            ws4.cell(row=r, column=col).number_format = CHF
        else:
            ws4.cell(row=r, column=col).number_format = NUM

    # Total col
    ws4.cell(row=r, column=15, value=yearly).fill = fill(bg_use)
    ws4.cell(row=r, column=15).font = font(bold=True)
    ws4.cell(row=r, column=15).alignment = right()
    ws4.cell(row=r, column=15).border = border_thin()
    if label in ("Ziel-Umsatz","Kosten (fix)","Betriebsergebnis"):
        ws4.cell(row=r, column=15).number_format = CHF
    else:
        ws4.cell(row=r, column=15).number_format = NUM

    # Avg col
    ws4.cell(row=r, column=16, value=mo_val).fill = fill(bg_use)
    ws4.cell(row=r, column=16).font = font(bold=is_highlight)
    ws4.cell(row=r, column=16).alignment = right()
    ws4.cell(row=r, column=16).border = border_thin()
    if label in ("Ziel-Umsatz","Kosten (fix)","Betriebsergebnis"):
        ws4.cell(row=r, column=16).number_format = CHF
    else:
        ws4.cell(row=r, column=16).number_format = NUM
    r += 1

r += 1
note(ws4, r, 2, "Planwerte. Tatsächlicher Umsatz hängt von Auslastung und Projekten ab. Werte exkl. MWST.")

# ── Freeze panes & print settings ────────────────────────────────────────────
for ws in [ws1, ws2, ws3, ws4]:
    ws.freeze_panes = "B5"
    ws.print_title_rows = "1:4"

path = "/home/user/yucker-studio/Yucker_Studio_Finanzen_2026.xlsx"
wb.save(path)
print(f"Gespeichert: {path}")
